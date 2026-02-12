#!/usr/bin/env python3
"""
D&D 5e 自动化人物卡解析器
支持全自动人物卡引擎（8分页，2400+公式）
完整还原Excel公式计算逻辑
"""

import json
import pandas as pd
from pathlib import Path
from typing import Dict, Any, Optional, List
import openpyxl
from openpyxl import load_workbook


class CharacterSheetEngine:
    """
    D&D 5e 自动化人物卡引擎
    完整还原Excel公式计算逻辑
    """
    
    # D&D 5e 标准职业熟练豁免映射
    CLASS_SAVE_PROFICIENCIES = {
        "战士": ["strength", "constitution"],
        "野蛮人": ["strength", "constitution"],
        "武僧": ["strength", "dexterity"],
        "游侠": ["strength", "dexterity"],
        "游荡者": ["dexterity", "intelligence"],
        "法师": ["intelligence", "wisdom"],
        "德鲁伊": ["intelligence", "wisdom"],
        "奇械师": ["intelligence", "constitution"],
        "邪术师": ["wisdom", "charisma"],
        "牧师": ["wisdom", "charisma"],
        "圣武士": ["wisdom", "charisma"],
        "吟游诗人": ["dexterity", "charisma"],
        "术士": ["constitution", "charisma"],
    }
    
    # D&D 5e 种族特性映射
    RACIAL_TRAITS = {
        "半兽人": {
            "ability_increase": {"strength": 2, "constitution": 1},
            "traits": ["darkvision", "relentless_endurance", "savage_attacks", "menacing"],
            "skill_proficiency": ["威吓"],
            "size": "medium",
            "speed": 30
        },
        "人类": {
            "ability_increase": {"all": 1},
            "traits": [],
            "size": "medium",
            "speed": 30
        },
        "精灵": {
            "ability_increase": {"dexterity": 2},
            "traits": ["darkvision", "keen_senses", "fey_ancestry", "trance"],
            "skill_proficiency": ["察觉"],
            "size": "medium",
            "speed": 30
        },
        "矮人": {
            "ability_increase": {"constitution": 2},
            "traits": ["darkvision", "dwarven_resilience", "stonecunning"],
            "size": "medium",
            "speed": 25
        },
        "半身人": {
            "ability_increase": {"dexterity": 2},
            "traits": ["lucky", "brave", "halfling_nimbleness"],
            "size": "small",
            "speed": 25
        },
        "龙裔": {
            "ability_increase": {"strength": 2, "charisma": 1},
            "traits": ["draconic_ancestry", "breath_weapon", "damage_resistance"],
            "size": "medium",
            "speed": 30
        },
        "侏儒": {
            "ability_increase": {"intelligence": 2},
            "traits": ["darkvision", "gnome_cunning"],
            "size": "small",
            "speed": 25
        },
        "半精灵": {
            "ability_increase": {"charisma": 2},
            "traits": ["darkvision", "fey_ancestry"],
            "size": "medium",
            "speed": 30
        },
        "提夫林": {
            "ability_increase": {"intelligence": 1, "charisma": 2},
            "traits": ["darkvision", "hellish_resistance", "infernal_legacy"],
            "size": "medium",
            "speed": 30
        }
    }
    
    # D&D 5e 战士战斗风格
    FIGHTING_STYLES = {
        "对决": {
            "name_en": "Dueling",
            "description": "单手持用一把近战武器，且并未同时持用其他武器时，伤害掷骰+2",
            "damage_bonus": 2,
            "condition": "单手单武器"
        },
        "双武器": {
            "name_en": "Two-Weapon Fighting",
            "description": "双武器战斗时，副手攻击可以加能力调整值",
            "damage_bonus": "ability_modifier",
            "condition": "双武器"
        },
        "箭术": {
            "name_en": "Archery",
            "description": "远程武器攻击检定+2",
            "attack_bonus": 2,
            "condition": "远程武器"
        },
        "防御": {
            "name_en": "Defense",
            "description": "着甲时AC+1",
            "ac_bonus": 1,
            "condition": "着甲"
        },
        "守护": {
            "name_en": "Protection",
            "description": "持盾时，可干扰5尺内被攻击的盟友",
            "condition": "持盾"
        },
        "重武器": {
            "name_en": "Great Weapon Fighting",
            "description": "重武器伤害骰1或2时可重掷",
            "condition": "重武器"
        }
    }
    
    # D&D 5e 战士职业特性
    FIGHTER_FEATURES = {
        1: ["战斗风格", "回气"],
        2: ["动作如潮"],
        3: ["武术范型"],
        5: ["额外攻击"],
        9: ["不屈"],
        11: ["额外攻击(2)"],
        13: ["不屈(2)"],
        17: ["不屈(3)", "动作如潮(2)"],
        20: ["额外攻击(3)"]
    }
    
    # 战士子职（武术范型）
    FIGHTER_SUBCLASSES = {
        "勇士": {
            "name_en": "Champion",
            "features": {
                3: ["强袭", "运动员"],
                7: ["额外战斗风格"],
                10: ["强袭+1"],
                15: ["超自然反射"],
                18: ["求生者"]
            }
        },
        "战斗大师": {
            "name_en": "Battle Master",
            "features": {
                3: ["卓越骰", "战法"],
                7: ["知己知彼"],
                10: ["精熟战法"],
                15: ["精熟战法+"],
                18: ["优势战法"]
            }
        },
        "奥法骑士": {
            "name_en": "Eldritch Knight",
            "features": {
                3: ["法术施放", "武器绑定"],
                7: ["战争魔法"],
                10: ["秘法打击"],
                15: ["战争魔法+"],
                18: ["精通战争魔法"]
            }
        }
    }
    
    # 技能与属性映射
    SKILL_ABILITY_MAP = {
        "运动": "strength",
        "体操": "dexterity",
        "巧手": "dexterity",
        "隐匿": "dexterity",
        "奥秘": "intelligence",
        "历史": "intelligence",
        "调查": "intelligence",
        "自然": "intelligence",
        "宗教": "intelligence",
        "驯兽": "wisdom",
        "洞悉": "wisdom",
        "医药": "wisdom",
        "察觉": "wisdom",
        "生存": "wisdom",
        "欺瞒": "charisma",
        "威吓": "charisma",
        "表演": "charisma",
        "游说": "charisma"
    }
    
    # 生命骰映射
    HIT_DICE = {
        "法师": "1d6", "术士": "1d6", "游荡者": "1d8",
        "吟游诗人": "1d8", "牧师": "1d8", "德鲁伊": "1d8",
        "邪术师": "1d8", "武僧": "1d8", "奇械师": "1d8",
        "战士": "1d10", "圣武士": "1d10", "游侠": "1d10",
        "野蛮人": "1d12"
    }
    
    def __init__(self):
        self.raw_data = {}
        self.calculated = {}
    
    def calculate_proficiency_bonus(self, level: int) -> int:
        """
        计算熟练加值
        =IF(等级<5,2,IF(等级<9,3,IF(等级<13,4,IF(等级<17,5,6))))
        """
        if level < 5:
            return 2
        elif level < 9:
            return 3
        elif level < 13:
            return 4
        elif level < 17:
            return 5
        else:
            return 6
    
    def calculate_ability_modifier(self, score: int) -> int:
        """
        计算属性调整值
        =INT(属性值/2-5)
        """
        return (score - 10) // 2
    
    def is_save_proficient(self, class_name: str, ability: str) -> bool:
        """
        判断职业是否熟练某项豁免
        =IF(OR(职业="战士",职业="武僧"...),"O","X")
        """
        profs = self.CLASS_SAVE_PROFICIENCIES.get(class_name, [])
        return ability in profs
    
    def calculate_save_bonus(self, ability_mod: int, is_proficient: bool, 
                            prof_bonus: int, other_bonus: int = 0) -> int:
        """
        计算豁免加值
        =修正值+IF(熟练="O",熟练加值,0)+其他加值
        """
        bonus = ability_mod
        if is_proficient:
            bonus += prof_bonus
        bonus += other_bonus
        return bonus
    
    def calculate_skill_bonus(self, ability_mod: int, is_proficient: bool,
                             prof_bonus: int, other_bonus: int = 0) -> int:
        """
        计算技能加值
        同豁免公式
        """
        return self.calculate_save_bonus(ability_mod, is_proficient, prof_bonus, other_bonus)
    
    def calculate_initiative(self, dex_mod: int, other_bonus: int = 0, 
                            has_alert: bool = False, is_bard_2: bool = False,
                            prof_bonus: int = 2) -> int:
        """
        计算先攻
        =敏捷修正+其他加值+IF(有警觉专长,5)+IF(吟游诗人2级+,INT(熟练加值/2))
        """
        init = dex_mod + other_bonus
        if has_alert:
            init += 5
        if is_bard_2:
            init += prof_bonus // 2
        return init


class AdvancedExcelParser(CharacterSheetEngine):
    """
    高级Excel人物卡解析器
    支持全自动8分页人物卡
    """
    
    def __init__(self):
        super().__init__()
        self.workbook = None
        self.sheets = {}
    
    def parse(self, file_path: str) -> Dict[str, Any]:
        """解析全自动人物卡"""
        try:
            self.workbook = load_workbook(file_path, data_only=True)
            self.sheets = {name: self.workbook[name] for name in self.workbook.sheetnames}
            
            character = {
                "basic": self._parse_basic_info(),
                "abilities": self._parse_abilities(),
                "combat": self._parse_combat_stats(),
                "saves": self._parse_saves(),
                "skills": self._parse_skills(),
                "equipment": self._parse_equipment(),
                "spells": self._parse_spellcasting(),
                "backpack": self._parse_backpack(),
                "features": self._parse_features(),
                "racial_traits": self._parse_racial_traits(),
                "class_features": self._parse_class_features(),
                "background": self._parse_background()
            }
            
            # 验证并补充计算
            character = self._validate_and_complete(character)
            
            return character
            
        except Exception as e:
            return {"error": f"解析失败: {str(e)}"}
    
    def _get_sheet_value(self, sheet_name: str, cell: str, default=None):
        """安全获取单元格值"""
        try:
            if sheet_name in self.sheets:
                val = self.sheets[sheet_name][cell].value
                return val if val is not None else default
        except:
            pass
        return default
    
    def _parse_basic_info(self) -> Dict[str, Any]:
        """解析基本信息（角色分页）"""
        sheet = "角色"
        
        # 尝试多种可能的单元格位置
        char_name = (self._get_sheet_value(sheet, "E3") or 
                    self._get_sheet_value(sheet, "B3") or
                    self._get_sheet_value("主要情况", "E3"))
        
        player_name = (self._get_sheet_value(sheet, "E4") or 
                      self._get_sheet_value(sheet, "B5"))
        
        # 种族在E6（根据扫描结果：B6是"种族"标签，E6是值）
        race = (self._get_sheet_value(sheet, "E6") or 
               self._get_sheet_value(sheet, "E5") or
               self._get_sheet_value("主要情况", "S4"))
        
        # 背景在E10（根据扫描结果：B10是"背景"标签，E10是值）
        background = (self._get_sheet_value(sheet, "E10") or 
                     self._get_sheet_value(sheet, "E6"))
        
        alignment = (self._get_sheet_value(sheet, "E7") or 
                    self._get_sheet_value(sheet, "A7"))
        
        # 职业：从主要情况!S3读取（"图表"分页引用职业）
        # 或者从W列的技能反推（如果运动+生存是战士/游侠/野蛮人）
        class_name = (self._get_sheet_value("主要情况", "S3") or
                     self._get_sheet_value("主要情况", "B2"))
        
        # 如果还找不到，根据技能反推
        if not class_name:
            skills_text = self._get_sheet_value(sheet, "W3", "")
            if "运动" in str(skills_text) and "求生" in str(skills_text):
                # 化外之民背景给运动+生存
                class_name = "战士"  # 默认，需要玩家确认
        
        # 等级从主要情况!W3读取
        level_val = (self._get_sheet_value("主要情况", "W3") or 
                    self._get_sheet_value(sheet, "W4"))
        
        try:
            level = int(level_val) if level_val else 1
        except:
            level = 1
        
        xp = self._get_sheet_value(sheet, "W5", 0)
        try:
            xp = int(xp) if xp else 0
        except:
            xp = 0
        
        return {
            "character_name": str(char_name) if char_name else "未命名角色",
            "player_name": str(player_name) if player_name else "",
            "race": str(race) if race else "",
            "class": str(class_name) if class_name else "",
            "level": level,
            "background": str(background) if background else "",
            "alignment": str(alignment) if alignment else "",
            "experience": xp
        }
    
    def _parse_abilities(self) -> Dict[str, Any]:
        """解析六维属性（主要情况分页）"""
        sheet = "主要情况"
        
        # E列是属性值，G列是修正值
        ability_map = {
            "strength": ("E8", "G8"),
            "dexterity": ("E10", "G10"),
            "constitution": ("E12", "G12"),
            "intelligence": ("E14", "G14"),
            "wisdom": ("E16", "G16"),
            "charisma": ("E18", "G18")
        }
        
        abilities = {}
        for ability, (score_cell, mod_cell) in ability_map.items():
            score = self._get_sheet_value(sheet, score_cell, 10)
            mod = self._get_sheet_value(sheet, mod_cell)
            
            try:
                score = int(score) if score else 10
            except:
                score = 10
            
            # 如果没有修正值，自动计算
            if mod is None:
                mod = self.calculate_ability_modifier(score)
            else:
                try:
                    mod = int(mod)
                except:
                    mod = self.calculate_ability_modifier(score)
            
            abilities[ability] = {
                "score": score,
                "modifier": mod
            }
        
        return abilities
    
    def _parse_combat_stats(self) -> Dict[str, Any]:
        """解析战斗数值"""
        sheet = "主要情况"
        
        level = self._get_sheet_value("角色", "W4", 1)
        try:
            level = int(level)
        except:
            level = 1
        
        prof_bonus = self.calculate_proficiency_bonus(level)
        
        # 获取各项数值
        max_hp = self._get_sheet_value(sheet, "AC12")
        current_hp = self._get_sheet_value(sheet, "AC8")
        ac = self._get_sheet_value(sheet, "S8")
        initiative = self._get_sheet_value(sheet, "P8")
        speed = self._get_sheet_value(sheet, "AC10")
        
        # 转换为整数
        try:
            max_hp = int(max_hp) if max_hp else 10
        except:
            max_hp = 10
        
        try:
            current_hp = int(current_hp) if current_hp else max_hp
        except:
            current_hp = max_hp
        
        try:
            ac = int(ac) if ac else 10
        except:
            ac = 10
        
        abilities = self._parse_abilities()
        dex_mod = abilities.get("dexterity", {}).get("modifier", 0)
        
        # 如果没有先攻值，自动计算
        if initiative is None:
            initiative = dex_mod
        else:
            try:
                initiative = int(initiative)
            except:
                initiative = dex_mod
        
        try:
            speed = int(speed) if speed else 30
        except:
            speed = 30
        
        # 生命骰
        class_name = self._parse_basic_info().get("class", "")
        hit_dice = self.HIT_DICE.get(class_name, "1d8")
        
        return {
            "max_hp": max_hp,
            "current_hp": current_hp,
            "temp_hp": 0,
            "ac": ac,
            "initiative": initiative,
            "speed": speed,
            "hit_dice": hit_dice,
            "proficiency_bonus": prof_bonus
        }
    
    def _parse_saves(self) -> Dict[str, Any]:
        """解析豁免检定"""
        sheet = "主要情况"
        
        basic = self._parse_basic_info()
        abilities = self._parse_abilities()
        prof_bonus = self.calculate_proficiency_bonus(basic.get("level", 1))
        class_name = basic.get("class", "")
        
        save_cells = {
            "strength": "K8",
            "dexterity": "K10",
            "constitution": "K12",
            "intelligence": "K14",
            "wisdom": "K16",
            "charisma": "K18"
        }
        
        saves = {}
        for ability, cell in save_cells.items():
            saved_value = self._get_sheet_value(sheet, cell)
            ability_mod = abilities.get(ability, {}).get("modifier", 0)
            is_prof = self.is_save_proficient(class_name, ability)
            
            if saved_value is not None:
                try:
                    bonus = int(saved_value)
                except:
                    bonus = self.calculate_save_bonus(ability_mod, is_prof, prof_bonus)
            else:
                bonus = self.calculate_save_bonus(ability_mod, is_prof, prof_bonus)
            
            saves[ability] = {
                "bonus": bonus,
                "proficient": is_prof
            }
        
        return saves
    
    def _parse_skills(self) -> Dict[str, Any]:
        """解析技能"""
        basic = self._parse_basic_info()
        abilities = self._parse_abilities()
        prof_bonus = self.calculate_proficiency_bonus(basic.get("level", 1))
        
        skills = {}
        for skill_name, ability in self.SKILL_ABILITY_MAP.items():
            ability_mod = abilities.get(ability, {}).get("modifier", 0)
            # 简化为非熟练（实际应从Excel读取B列熟练标记）
            bonus = ability_mod
            
            skills[skill_name] = {
                "ability": ability,
                "bonus": bonus,
                "proficient": False
            }
        
        return skills
    
    def _parse_equipment(self) -> Dict[str, Any]:
        """解析装备"""
        sheet = "背包"
        
        weapons = []
        armor = ""
        shield = False
        items = []
        
        # 读取武器和装备
        for row in range(5, 20):
            item_name = self._get_sheet_value(sheet, f"B{row}")
            if item_name:
                item_str = str(item_name)
                items.append(item_str)
                
                # 简单判断武器/护甲
                if any(w in item_str for w in ["剑", "斧", "锤", "弓", "弩", "矛", "杖", "匕首", "镰", "锏"]):
                    weapons.append(item_str)
                elif any(a in item_str for a in ["甲", "盔", "盾", "铠"]):
                    if "盾" in item_str:
                        shield = True
                    else:
                        armor = item_str
        
        return {
            "weapons": weapons,
            "armor": armor,
            "shield": shield,
            "items": items
        }
    
    def _parse_backpack(self) -> Dict[str, Any]:
        """解析背包详情"""
        sheet = "背包"
        
        total_weight = self._get_sheet_value(sheet, "V3", 0)
        total_cost = self._get_sheet_value(sheet, "AS3", 0)
        
        try:
            total_weight = float(total_weight) if total_weight else 0
        except:
            total_weight = 0
        
        try:
            total_cost = float(total_cost) if total_cost else 0
        except:
            total_cost = 0
        
        # 详细装备列表
        equipment_list = []
        for row in range(5, 30):
            name = self._get_sheet_value(sheet, f"B{row}")
            qty = self._get_sheet_value(sheet, f"S{row}", 1)
            weight = self._get_sheet_value(sheet, f"V{row}", 0)
            
            if name:
                try:
                    qty = int(qty) if qty else 1
                except:
                    qty = 1
                
                try:
                    weight = float(weight) if weight else 0
                except:
                    weight = 0
                
                equipment_list.append({
                    "name": str(name),
                    "quantity": qty,
                    "weight": weight
                })
        
        return {
            "total_weight": total_weight,
            "total_cost": total_cost,
            "items": equipment_list
        }
    
    def _parse_spellcasting(self) -> Dict[str, Any]:
        """解析施法能力"""
        sheet = "施法"
        
        spell_ability = self._get_sheet_value(sheet, "H2")
        save_dc = self._get_sheet_value(sheet, "O2")
        attack_bonus = self._get_sheet_value(sheet, "V2")
        cantrips_known = self._get_sheet_value(sheet, "Y2")
        
        # 法术位
        spell_slots = {}
        for i, cell in enumerate(["AC2", "AC3", "AC4", "AC5", "AC6", "AC7", "AC8", "AC9", "AC10"], 1):
            slots = self._get_sheet_value(sheet, cell)
            try:
                spell_slots[f"{i}st" if i == 1 else f"{i}nd" if i == 2 else f"{i}rd" if i == 3 else f"{i}th"] = int(slots) if slots else 0
            except:
                spell_slots[f"{i}st" if i == 1 else f"{i}nd" if i == 2 else f"{i}rd" if i == 3 else f"{i}th"] = 0
        
        # 从法术大全读取已知法术
        known_spells = []
        if "法术大全" in self.sheets:
            spells_sheet = self.sheets["法术大全"]
            for row in range(1, 50):
                spell_name = spells_sheet.cell(row=row, column=2).value
                if spell_name and spell_name != "法术名":
                    known_spells.append(str(spell_name))
        
        if not spell_ability and not save_dc:
            return {}  # 非施法职业
        
        return {
            "spellcasting_ability": str(spell_ability) if spell_ability else "",
            "spell_save_dc": int(save_dc) if save_dc else 0,
            "spell_attack_bonus": int(attack_bonus) if attack_bonus else 0,
            "cantrips_known": str(cantrips_known) if cantrips_known else "",
            "spell_slots": spell_slots,
            "known_spells": known_spells
        }
    
    def _parse_features(self) -> Dict[str, Any]:
        """解析职业和种族特性"""
        features = {
            "class_features": [],
            "racial_traits": [],
            "feats": [],
            "fighting_style": None,
            "subclass": None
        }
        
        # 从主要情况分页读取特性
        if "主要情况" in self.sheets:
            sheet = self.sheets["主要情况"]
            # AV列是职业特性
            for row in range(2, 30):
                feature = sheet.cell(row=row, column=48).value  # AV列
                if feature:
                    features["class_features"].append(str(feature))
            
            # BM列是种族特性
            for row in range(2, 20):
                trait = sheet.cell(row=row, column=65).value  # BM列
                if trait:
                    features["racial_traits"].append(str(trait))
        
        # 读取战斗风格（从BB列或AW列）
        if "主要情况" in self.sheets:
            sheet = self.sheets["主要情况"]
            # AW3 是战斗风格标题，BB4 是具体选择
            fighting_style_cell = sheet["BB4"].value
            if fighting_style_cell and "对决" in str(fighting_style_cell):
                features["fighting_style"] = "对决"
            elif fighting_style_cell and "双武器" in str(fighting_style_cell):
                features["fighting_style"] = "双武器"
            elif fighting_style_cell and "箭术" in str(fighting_style_cell):
                features["fighting_style"] = "箭术"
            elif fighting_style_cell and "防御" in str(fighting_style_cell):
                features["fighting_style"] = "防御"
            elif fighting_style_cell and "守护" in str(fighting_style_cell):
                features["fighting_style"] = "守护"
            elif fighting_style_cell and "重武器" in str(fighting_style_cell):
                features["fighting_style"] = "重武器"
        
        return features
    
    def _parse_racial_traits(self) -> Dict[str, Any]:
        """解析种族特性"""
        basic = self._parse_basic_info()
        race = basic.get("race", "")
        
        traits = {
            "race_name": race,
            "traits": [],
            "ability_increase": {},
            "skill_proficiencies": [],
            "darkvision": 0,
            "speed": 30,
            "size": "medium"
        }
        
        # 从RACIAL_TRAITS映射获取
        race_data = self.RACIAL_TRAITS.get(race, {})
        if race_data:
            traits["ability_increase"] = race_data.get("ability_increase", {})
            traits["traits"] = race_data.get("traits", [])
            traits["skill_proficiencies"] = race_data.get("skill_proficiency", [])
            traits["speed"] = race_data.get("speed", 30)
            traits["size"] = race_data.get("size", "medium")
            
            if "darkvision" in race_data.get("traits", []):
                traits["darkvision"] = 60
        
        # 特殊处理半兽人
        if "半兽人" in race:
            traits["traits"] = [
                {
                    "name": "黑暗视觉",
                    "name_en": "Darkvision",
                    "description": "60尺黑暗视觉"
                },
                {
                    "name": "不屈",
                    "name_en": "Relentless Endurance",
                    "description": "HP降至0但未直接致死时，改为降至1HP（1次/长休）",
                    "frequency": "1次/长休",
                    "available": True
                },
                {
                    "name": "凶恶攻击",
                    "name_en": "Savage Attacks",
                    "description": "近战武器暴击时，额外追加一个伤害骰",
                    "trigger": "近战暴击"
                },
                {
                    "name": "凶恶外观",
                    "name_en": "Menacing",
                    "description": "威吓技能熟练",
                    "grants_proficiency": ["威吓"]
                }
            ]
            traits["darkvision"] = 60
            traits["skill_proficiencies"].append("威吓")
        
        return traits
    
    def _parse_class_features(self) -> Dict[str, Any]:
        """解析职业特性"""
        basic = self._parse_basic_info()
        class_name = basic.get("class", "")
        level = basic.get("level", 1)
        
        features = {
            "class_name": class_name,
            "level": level,
            "features_available": [],
            "features_by_level": {},
            "hit_dice": self.HIT_DICE.get(class_name, "1d8")
        }
        
        # 战士特性
        if "战士" in class_name:
            fighter_features = {
                "second_wind": {
                    "name": "回气",
                    "name_en": "Second Wind",
                    "level": 1,
                    "action": "附赠动作",
                    "description": f"恢复1d10+{level} HP",
                    "frequency": "1次/短休或长休",
                    "formula": f"1d10+{level}",
                    "available": level >= 1
                },
                "fighting_style": {
                    "name": "战斗风格",
                    "name_en": "Fighting Style",
                    "level": 1,
                    "description": "选择一种战斗风格",
                    "available": level >= 1
                },
                "action_surge": {
                    "name": "动作如潮",
                    "name_en": "Action Surge",
                    "level": 2,
                    "description": "自己回合内多一个动作",
                    "frequency": "1次/短休或长休" if level < 17 else "2次/短休或长休",
                    "available": level >= 2
                },
                "martial_archetype": {
                    "name": "武术范型",
                    "name_en": "Martial Archetype",
                    "level": 3,
                    "description": "选择子职：勇士、战斗大师、奥法骑士",
                    "available": level >= 3
                },
                "extra_attack": {
                    "name": "额外攻击",
                    "name_en": "Extra Attack",
                    "level": 5,
                    "description": "攻击动作可攻击多次",
                    "attacks": 2 if level < 11 else (3 if level < 20 else 4),
                    "available": level >= 5
                },
                "indomitable": {
                    "name": "不屈",
                    "name_en": "Indomitable",
                    "level": 9,
                    "description": "重掷失败的豁免检定",
                    "frequency": "1次/长休" if level < 13 else ("2次/长休" if level < 17 else "3次/长休"),
                    "available": level >= 9
                }
            }
            
            features["features_available"] = [
                f for f in fighter_features.values() if f["available"]
            ]
            features["features_by_level"] = fighter_features
            
            # 读取已选择的战斗风格
            fighting_style = self._parse_features().get("fighting_style")
            if fighting_style:
                style_data = self.FIGHTING_STYLES.get(fighting_style, {})
                features["selected_fighting_style"] = {
                    "name": fighting_style,
                    "name_en": style_data.get("name_en", ""),
                    "description": style_data.get("description", ""),
                    "damage_bonus": style_data.get("damage_bonus", 0),
                    "condition": style_data.get("condition", "")
                }
        
        return features
    
    def _parse_background(self) -> Dict[str, Any]:
        """解析背景信息"""
        sheet = "角色"
        
        background = {
            "name": self._get_sheet_value(sheet, "E10"),
            "personality_trait": self._get_sheet_value(sheet, "W9"),
            "ideal": self._get_sheet_value(sheet, "W16"),
            "bond": self._get_sheet_value(sheet, "W17"),
            "flaw": self._get_sheet_value(sheet, "W18"),
            "story": self._get_sheet_value(sheet, "W20"),
            "skills": [],
            "tools": [],
            "languages": [],
            "equipment": [],
            "feature": None
        }
        
        # 读取W列的技能、工具、语言、装备
        skills_text = self._get_sheet_value(sheet, "W3")
        if skills_text:
            background["skills"] = [s.strip() for s in str(skills_text).split("，")]
        
        tools_text = self._get_sheet_value(sheet, "W4")
        if tools_text:
            background["tools"] = [t.strip() for t in str(tools_text).split("，")]
        
        languages_text = self._get_sheet_value(sheet, "W5")
        if languages_text:
            background["languages"] = [l.strip() for l in str(languages_text).split("，")]
        
        equipment_text = self._get_sheet_value(sheet, "W6")
        if equipment_text:
            background["equipment"] = [e.strip() for e in str(equipment_text).split("，")]
        
        # 背景特性
        bg_name = background["name"]
        if bg_name and "化外之民" in str(bg_name):
            background["feature"] = {
                "name": "行者无疆",
                "name_en": "Wanderer",
                "description": "你拥有极佳的记忆力，能够回忆地形、定居点、水源等。你能在野外找到食物和水。"
            }
        
        return background
    
    def _validate_and_complete(self, character: Dict[str, Any]) -> Dict[str, Any]:
        """验证数据完整性并补充计算"""
        
        # 确保基础结构
        if "abilities" not in character:
            character["abilities"] = {}
        
        # 如果Excel没有计算出先攻，手动计算
        if "combat" in character:
            combat = character["combat"]
            abilities = character.get("abilities", {})
            
            if combat.get("initiative") is None or combat.get("initiative") == 0:
                dex_mod = abilities.get("dexterity", {}).get("modifier", 0)
                combat["initiative"] = dex_mod
            
            # 确保current_hp不大于max_hp
            if combat.get("current_hp", 0) > combat.get("max_hp", 0):
                combat["current_hp"] = combat["max_hp"]
        
        return character


def parse_character(file_path: str) -> Dict[str, Any]:
    """
    解析人物卡（自动检测格式）
    
    Args:
        file_path: 人物卡文件路径
        
    Returns:
        解析后的角色数据，包含自动计算的修正值和豁免
    """
    path = Path(file_path)
    
    if not path.exists():
        return {"error": f"文件不存在: {file_path}"}
    
    suffix = path.suffix.lower()
    
    if suffix in ['.xlsx', '.xls']:
        parser = AdvancedExcelParser()
        return parser.parse(file_path)
    elif suffix == '.json':
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                character = json.load(f)
            
            # 对JSON数据也进行自动计算和验证
            character = _process_character_data(character)
            return character
            
        except Exception as e:
            return {"error": f"JSON解析失败: {str(e)}"}
    else:
        return {"error": f"不支持的文件格式: {suffix}"}


def _calculate_modifier(score: int) -> int:
    """根据属性值计算调整值"""
    return (score - 10) // 2


def _process_character_data(character: Dict[str, Any]) -> Dict[str, Any]:
    """
    处理角色数据，自动计算调整值、豁免等
    """
    # 确保基础结构
    if "abilities" not in character:
        character["abilities"] = {}
    
    abilities = character["abilities"]
    
    # 计算各属性调整值
    ability_names = ["str", "dex", "con", "int", "wis", "cha"]
    ability_full_names = {
        "str": "strength",
        "dex": "dexterity", 
        "con": "constitution",
        "int": "intelligence",
        "wis": "wisdom",
        "cha": "charisma"
    }
    
    # 计算调整值
    for short_name, full_name in ability_full_names.items():
        score = abilities.get(short_name) or abilities.get(full_name, 10)
        modifier = _calculate_modifier(score)
        abilities[f"{short_name}_modifier"] = modifier
        # 同时保存到全称键
        abilities[full_name] = score
    
    # 获取等级和熟练加值
    level = character.get("basic", {}).get("level", 1)
    proficiency_bonus = (level - 1) // 4 + 2  # 1-4级+2, 5-8级+3, etc.
    
    if "combat" not in character:
        character["combat"] = {}
    combat = character["combat"]
    combat["proficiency"] = combat.get("proficiency", proficiency_bonus)
    
    # 计算先攻（如果没有）
    if combat.get("initiative") is None or combat.get("initiative") == 0:
        combat["initiative"] = abilities.get("dex_modifier", 0)
    
    # 计算豁免
    if "saves" not in character:
        character["saves"] = {}
    
    # 战士的熟练豁免是力量和体质
    class_name = character.get("basic", {}).get("class", "").lower()
    default_save_proficiency = []
    if "fighter" in class_name or "战士" in class_name:
        default_save_proficiency = ["str", "con"]
    elif "wizard" in class_name or "法师" in class_name:
        default_save_proficiency = ["int", "wis"]
    elif "rogue" in class_name or "游荡" in class_name:
        default_save_proficiency = ["dex", "int"]
    # 可以添加更多职业...
    
    saves = character["saves"]
    for short_name in ability_names:
        modifier = abilities.get(f"{short_name}_modifier", 0)
        is_proficient = short_name in default_save_proficiency
        
        if short_name not in saves:
            saves[short_name] = {}
        
        save_bonus = modifier + (proficiency_bonus if is_proficient else 0)
        saves[short_name] = {
            "modifier": modifier,
            "proficient": is_proficient,
            "bonus": save_bonus
        }
    
    # 计算技能（简化版）
    if "skills" not in character:
        character["skills"] = {}
    
    # 技能对应的属性
    skill_abilities = {
        "acrobatics": "dex", "sleight_of_hand": "dex", "stealth": "dex",
        "arcana": "int", "history": "int", "investigation": "int", "nature": "int", "religion": "int",
        "animal_handling": "wis", "insight": "wis", "medicine": "wis", "perception": "wis", "survival": "wis",
        "deception": "cha", "intimidation": "cha", "performance": "cha", "persuasion": "cha",
        "athletics": "str"
    }
    
    prof_skills = character.get("skills_proficiency", [])
    skills = character["skills"]
    
    for skill_name, ability in skill_abilities.items():
        ability_mod = abilities.get(f"{ability}_modifier", 0)
        is_proficient = skill_name in prof_skills
        bonus = ability_mod + (proficiency_bonus if is_proficient else 0)
        
        skills[skill_name] = {
            "ability": ability,
            "modifier": ability_mod,
            "proficient": is_proficient,
            "bonus": bonus
        }
    
    # 自动计算Max HP（如果未设置或需要重新计算）
    if combat.get("max_hp", 0) == 0:
        # 职业生命骰映射
        hit_dice = {
            "fighter": 10, "战士": 10,
            "wizard": 6, "法师": 6,
            "rogue": 8, "游荡": 8, "游荡者": 8,
            "cleric": 8, "牧师": 8,
            "barbarian": 12, "野蛮人": 12,
            "bard": 8, "吟游诗人": 8,
            "druid": 8, "德鲁伊": 8,
            "monk": 8, "武僧": 8,
            "paladin": 10, "圣武士": 10, "圣骑士": 10,
            "ranger": 10, "游侠": 10,
            "sorcerer": 6, "术士": 6,
            "warlock": 8, "邪术师": 8,
        }
        
        # 获取职业生命骰
        hd = 8  # 默认
        for cls, dice in hit_dice.items():
            if cls in class_name:
                hd = dice
                break
        
        # 计算体质调整值
        con_score = abilities.get("con", abilities.get("constitution", 10))
        con_mod = _calculate_modifier(con_score)
        
        # 计算Max HP
        # 1级: 生命骰最大值 + 体质调整值
        # 后续等级: (生命骰平均值向上取整) + 体质调整值
        # level 已在前面定义
        first_level_hp = hd + con_mod
        
        if level > 1:
            # 后续等级平均HP (向上取整)
            avg_hp_per_level = (hd // 2) + 1
            additional_hp = (avg_hp_per_level + con_mod) * (level - 1)
            max_hp = first_level_hp + additional_hp
        else:
            max_hp = first_level_hp
        
        combat["max_hp"] = max_hp
        combat["hit_dice"] = f"{level}d{hd}"
    
    # 确保current_hp不大于max_hp
    if combat.get("current_hp", 0) > combat.get("max_hp", 0):
        combat["current_hp"] = combat["max_hp"]
    if combat.get("current_hp", 0) == 0 and combat.get("max_hp", 0) > 0:
        combat["current_hp"] = combat["max_hp"]
    
    return character


def get_character_summary(character: Dict[str, Any]) -> str:
    """生成角色摘要"""
    if "error" in character:
        return f"❌ {character['error']}"
    
    basic = character.get("basic", {})
    combat = character.get("combat", {})
    abilities = character.get("abilities", {})
    saves = character.get("saves", {})
    
    # 构建豁免显示
    save_strs = []
    for ability in ["strength", "dexterity", "constitution", "intelligence", "wisdom", "charisma"]:
        save_data = saves.get(ability, {})
        bonus = save_data.get("bonus", 0)
        prof = "★" if save_data.get("proficient") else " "
        abbr = ability[:3].upper()
        save_strs.append(f"{prof}{abbr}{bonus:+d}")
    
    summary = f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
🎭 {basic.get('character_name', '未知角色')}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
📋 {basic.get('race', '')} {basic.get('class', '')} | 等级 {basic.get('level', 1)}
   玩家: {basic.get('player_name', '未知')} | 阵营: {basic.get('alignment', '')}

❤️ HP: {combat.get('current_hp', 0)}/{combat.get('max_hp', 0)} | 🛡️ AC: {combat.get('ac', 10)}
⚡ 先攻: {combat.get('initiative', 0):+d} | 🏃 速度: {combat.get('speed', 30)}尺 | 🎲 {combat.get('hit_dice', '1d8')}

【六维属性】
   力量 {abilities.get('strength', {}).get('score', 10):2d} ({abilities.get('strength', {}).get('modifier', 0):+d})
   敏捷 {abilities.get('dexterity', {}).get('score', 10):2d} ({abilities.get('dexterity', {}).get('modifier', 0):+d})
   体质 {abilities.get('constitution', {}).get('score', 10):2d} ({abilities.get('constitution', {}).get('modifier', 0):+d})
   智力 {abilities.get('intelligence', {}).get('score', 10):2d} ({abilities.get('intelligence', {}).get('modifier', 0):+d})
   感知 {abilities.get('wisdom', {}).get('score', 10):2d} ({abilities.get('wisdom', {}).get('modifier', 0):+d})
   魅力 {abilities.get('charisma', {}).get('score', 10):2d} ({abilities.get('charisma', {}).get('modifier', 0):+d})

【豁免检定】{' | '.join(save_strs[:3])}
           {' | '.join(save_strs[3:])}
"""
    
    # 添加种族特性
    racial = character.get("racial_traits", {})
    if racial and racial.get("traits"):
        summary += f"""
【种族特性 - {racial.get('race_name', '')}】
"""
        for trait in racial.get("traits", []):
            if isinstance(trait, dict):
                name = trait.get("name", "")
                desc = trait.get("description", "")
                if name:
                    summary += f"   □ {name}: {desc[:40]}\n"
        if racial.get("darkvision", 0) > 0:
            summary += f"   👁️ 黑暗视觉: {racial['darkvision']}尺\n"
    
    # 添加职业特性
    class_data = character.get("class_features", {})
    if class_data and class_data.get("features_available"):
        summary += f"""
【职业特性 - {class_data.get('class_name', '')}】
"""
        # 显示可用的特性
        for feature in class_data.get("features_available", []):
            if isinstance(feature, dict):
                name = feature.get("name", "")
                action = feature.get("action", "")
                desc = feature.get("description", "")
                freq = feature.get("frequency", "")
                if name:
                    action_str = f"[{action}]" if action else ""
                    freq_str = f" ({freq})" if freq else ""
                    summary += f"   □ {name}{action_str}: {desc[:35]}{freq_str}\n"
        
        # 显示战斗风格
        selected_style = class_data.get("selected_fighting_style")
        if selected_style:
            summary += f"""
   ⚔️ 战斗风格: {selected_style.get('name', '')}
      {selected_style.get('description', '')[:50]}
"""
    
    # 添加背景信息
    bg = character.get("background", {})
    if bg and bg.get("name"):
        summary += f"""
【背景 - {bg.get('name', '')}】
   性格: {str(bg.get('personality_trait', ''))[:30]}
   理想: {str(bg.get('ideal', ''))[:30]}
"""
        if bg.get("feature"):
            feat = bg.get("feature", {})
            summary += f"   特性: {feat.get('name', '')}\n"
    
    # 添加施法信息
    spells = character.get("spells", {})
    if spells:
        summary += f"""
【施法能力】
   关键属性: {spells.get('spellcasting_ability', '')}
   法术豁免DC: {spells.get('spell_save_dc', 0)} | 法术攻击: {spells.get('spell_attack_bonus', 0):+d}
"""
    
    # 添加装备信息
    equipment = character.get("equipment", {})
    if equipment.get("weapons"):
        summary += f"""
【装备】
   武器: {', '.join(equipment.get('weapons', []))}
   护甲: {equipment.get('armor', '无')} {'+ 盾牌' if equipment.get('shield') else ''}
"""
    
    summary += "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
    
    return summary


def export_to_json(character: Dict[str, Any], output_path: str):
    """导出为JSON格式"""
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(character, f, indent=2, ensure_ascii=False)


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        character = parse_character(file_path)
        
        if "error" not in character:
            print(json.dumps(character, indent=2, ensure_ascii=False))
            print("\n" + get_character_summary(character))
            
            # 可选：导出为JSON
            if len(sys.argv) > 2:
                export_to_json(character, sys.argv[2])
                print(f"\n✅ 已导出至: {sys.argv[2]}")
        else:
            print(f"❌ {character['error']}")
    else:
        print("用法: python character_parser.py <人物卡.xlsx> [输出.json]")
